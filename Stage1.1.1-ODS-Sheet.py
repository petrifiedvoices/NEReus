#!/usr/bin/env python3
"""
POS Tagging Gold Standard Creator with Line-Based Highlighting
Generates XLSX spreadsheet for Latin inscription POS tagging validation
Shows inscription lines with matching characters highlighted
"""

import json
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

# Universal Dependencies v2 POS tags
UD_POS_TAGS = [
    "ADJ",   # adjective
    "ADP",   # adposition (preposition/postposition)
    "ADV",   # adverb
    "AUX",   # auxiliary verb
    "CCONJ", # coordinating conjunction
    "DET",   # determiner
    "INTJ",  # interjection
    "NOUN",  # noun
    "NUM",   # numeral
    "PART",  # particle
    "PRON",  # pronoun
    "PROPN", # proper noun
    "PUNCT", # punctuation
    "SCONJ", # subordinating conjunction
    "SYM",   # symbol
    "VERB",  # verb
    "X"      # other (foreign words, typos, abbreviations)
]

# Instructions for the annotator (will go in column)
INSTRUCTIONS = """Universal Dependencies v2 POS Tags:
ADJ=adjective (magnus, bonus)
ADP=preposition/postposition (in, ad, cum)
ADV=adverb (bene, semper, non)
AUX=auxiliary verb (sum as copula/auxiliary)
CCONJ=coordinating conjunction (et, -que, aut)
DET=determiner (hic, ille, ipse)
INTJ=interjection (o, eheu)
NOUN=noun (homo, res, urbs)
NUM=numeral (unus, tres, XX)
PART=particle (ne, -ne interrogative)
PRON=pronoun (ego, qui, is)
PROPN=proper noun (Roma, Iulius, Marcus)
PUNCT=punctuation (. , :)
SCONJ=subordinating conjunction (ut, cum, si)
SYM=symbol (special symbols)
VERB=verb (amo, facio, dico)
X=other (foreign, abbreviations, uncertain)"""


def clean_for_matching(text):
    """Remove epigraphic notation for character matching"""
    # Remove brackets, parentheses, question marks, numbers, and spaces
    cleaned = re.sub(r'[\[\]()?\d\s/]', '', text)
    return cleaned.lower()


def find_matching_line(word, inscription_lines):
    """
    Find which line contains the best matching substring for the word
    Returns the line index and the line text
    """
    if not word or not inscription_lines:
        return 0, inscription_lines[0] if inscription_lines else ""
    
    best_line_idx = 0
    best_score = 0
    
    for idx, line in enumerate(inscription_lines):
        # Find the best matching substring in this line
        start, end = find_matching_substring(line, word)
        
        if start != -1:
            # Score based on how much of the word was matched
            score = len(word) / max(1, (end - start))  # Prefer compact matches
            
            if score > best_score:
                best_score = score
                best_line_idx = idx
    
    return best_line_idx, inscription_lines[best_line_idx] if inscription_lines else ""


def find_matching_substring(line_text, word):
    """
    Find the substring in line_text that best matches the word
    Returns start and end positions of the match
    """
    word_lower = word.lower()
    best_start = -1
    best_end = -1
    best_score = 0
    
    # Try to find the word's characters in sequence
    for start_pos in range(len(line_text)):
        word_idx = 0
        end_pos = start_pos
        
        while end_pos < len(line_text) and word_idx < len(word_lower):
            char = line_text[end_pos].lower()
            
            # If this character matches the next character we need
            if char == word_lower[word_idx]:
                word_idx += 1
            # Skip epigraphic notation and non-letter characters
            elif not char.isalpha():
                pass  # Just move forward in the line
            else:
                # This is a letter that doesn't match - stop this attempt
                break
            
            end_pos += 1
        
        # If we matched the whole word
        if word_idx == len(word_lower):
            score = word_idx / max(1, (end_pos - start_pos))  # Prefer compact matches
            if score > best_score:
                best_score = score
                best_start = start_pos
                best_end = end_pos
    
    return best_start, best_end


def create_highlighted_text(line_text, word):
    """
    Create rich text with matching substring bolded
    Returns a CellRichText object for openpyxl
    """
    if not word or not line_text:
        return line_text
    
    # Find the matching substring
    start, end = find_matching_substring(line_text, word)
    
    if start == -1:
        # No match found, return plain text
        return line_text
    
    # Create rich text with the matching part bolded
    result_blocks = []
    
    # Add text before the match (not bold)
    if start > 0:
        result_blocks.append(TextBlock(InlineFont(), line_text[:start]))
    
    # Add the matched substring (bold)
    result_blocks.append(TextBlock(InlineFont(b=True), line_text[start:end]))
    
    # Add text after the match (not bold)
    if end < len(line_text):
        result_blocks.append(TextBlock(InlineFont(), line_text[end:]))
    
    # Return as CellRichText
    if result_blocks:
        return CellRichText(result_blocks)
    else:
        return line_text


def create_validation_spreadsheet(json_file, output_file):
    """Create XLSX spreadsheet with line-based highlighting for POS tagging"""
    
    # Load data
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Process first 20 inscriptions
    for idx, inscription_data in enumerate(data[:20]):
        list_id = inscription_data.get('LIST-ID', f'Unknown_{idx}')
        inscription_text = inscription_data.get('inscription', '')
        interpretive_text = inscription_data.get('text_interpretive_word', '')
        inscription_type = inscription_data.get('type_of_inscription_auto', 'unknown')
        
        # Skip if no interpretive text
        if not interpretive_text:
            continue
        
        # Split inscription into lines
        inscription_lines = inscription_text.split('/')
        
        # Create sheet named by LIST-ID
        sheet_name = str(list_id)[:31]  # Excel sheet name limit
        ws = wb.create_sheet(title=sheet_name)
        
        # Add headers
        headers = [
            'Inscription_Line', 'Interpretive_Word', 'POS', 'Notes',
            'Instructions', 'Full_Inscription', 'Full_Interpretive', 
            'Type_of_Inscription', 'LIST_ID'
        ]
        ws.append(headers)
        
        # Split interpretive text into words
        interpretive_words = interpretive_text.split()
        
        # Create POS tag validation
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(UD_POS_TAGS)}"',
            allow_blank=True
        )
        dv.error = 'Please select a valid UD POS tag'
        dv.errorTitle = 'Invalid POS Tag'
        
        # Track which line we're currently on for sequential matching
        current_line_idx = 0
        
        # Add data rows
        for row_idx, word in enumerate(interpretive_words, start=2):
            # Find the line containing this word's characters
            line_idx, line_text = find_matching_line(word, inscription_lines)
            
            # Update current line index if we've moved forward
            if line_idx >= current_line_idx:
                current_line_idx = line_idx
            
            # Create highlighted version of the line
            highlighted_line = create_highlighted_text(line_text, word)
            
            row_data = [
                highlighted_line,   # Inscription_Line (with highlighting)
                word,              # Interpretive_Word
                '',                # POS (empty for annotation)
                '',                # Notes (empty for annotation)
                INSTRUCTIONS if row_idx == 2 else '',  # Instructions (only first row)
                inscription_text if row_idx == 2 else '',  # Full inscription (only first row)
                interpretive_text if row_idx == 2 else '',  # Full interpretive (only first row)
                inscription_type if row_idx == 2 else '',  # Type (only first row)
                str(list_id) if row_idx == 2 else ''  # LIST_ID (only first row)
            ]
            
            # Add the row
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Add validation to POS column (column C)
            dv.add(f'C{row_idx}')
        
        ws.add_data_validation(dv)
        
        # Adjust column widths for readability
        ws.column_dimensions['A'].width = 60  # Inscription_Line
        ws.column_dimensions['B'].width = 20  # Interpretive_Word
        ws.column_dimensions['C'].width = 12  # POS
        ws.column_dimensions['D'].width = 30  # Notes
        ws.column_dimensions['E'].width = 40  # Instructions
        ws.column_dimensions['F'].width = 50  # Full_Inscription
        ws.column_dimensions['G'].width = 50  # Full_Interpretive
        ws.column_dimensions['H'].width = 20  # Type_of_Inscription
        ws.column_dimensions['I'].width = 12  # LIST_ID
        
        # Freeze the header row
        ws.freeze_panes = 'A2'
        
        print(f"Processed inscription {list_id}: {len(interpretive_words)} tokens")
    
    # Save as XLSX (LibreOffice can open this)
    wb.save(output_file)
    print(f"\nValidation spreadsheet saved to: {output_file}")
    
    # Print summary
    sheet_count = len(wb.sheetnames)
    print(f"Created {sheet_count} sheets for annotation")
    print("\nColumns structure:")
    print("- Column A: Inscription line with matching characters bolded")
    print("- Column B: Clean word to tag")
    print("- Column C: POS tag (use dropdown)")
    print("- Column D: Notes for difficult cases")
    print("- Columns E-I: Metadata and instructions")
    print("\nAnnotators should focus on filling column C (POS) using the dropdown options")


if __name__ == "__main__":
    # Example usage
    input_file = "POS-LIST-test1.json"
    output_file = datetime.now().strftime("%Y%m%d") + "-POS-GoldStandard.xlsx"
    
    create_validation_spreadsheet(input_file, output_file)
    
    print("\n" + "="*50)
    print("ANNOTATION WORKFLOW:")
    print("1. Open the .xlsx file in LibreOffice Calc")
    print("2. Each sheet represents one inscription")
    print("3. Review the Interpretive_Word column (B)")
    print("4. Look at the inscription line in column A (bolded chars match the word)")
    print("5. Select appropriate POS tag from dropdown in column C")
    print("6. Add any notes about difficult cases in column D")
    print("7. Save regularly to preserve your work")
    print("="*50)